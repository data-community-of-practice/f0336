"""
Crossref + OpenAlex Affiliation Fetch
=======================================
Reads the unique authors Excel file (output of extract_unique_authors.py)
and fetches each author's affiliated organisation(s).

Strategy:
  1. Query Crossref for each DOI and extract author affiliations.
  2. For authors still missing affiliations, fall back to OpenAlex
     (which aggregates Crossref, ORCID, PubMed, MAG, and publisher data).
  3. Report which source provided each affiliation.

Resilience features:
  - JSON caches for Crossref and OpenAlex (resumable)
  - Periodic saves
  - Retry with backoff
  - Graceful Ctrl+C handling

Setup:
  1. Place config.ini in the SAME folder as this script with:
       [crossref]
       email = yourreal@email.com
       delay = 1
       save_every = 50
       max_retries = 3
  2. pip install openpyxl requests

Usage:
  python fetch_affiliations.py <unique_authors.xlsx> [output.xlsx] [config.ini]

  From Spyder:
  !python "path/to/fetch_affiliations.py" "path/to/unique_authors.xlsx"
"""

import sys
import re
import json
import time
import configparser
from pathlib import Path
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

CROSSREF_API = "https://api.crossref.org/works/"
OPENALEX_API = "https://api.openalex.org/works/doi:"
SCRIPT_DIR = Path(__file__).resolve().parent


# ── Config ──────────────────────────────────────────────────────────────

def load_config(config_path=None):
    if config_path is None:
        config_path = SCRIPT_DIR / "config.ini"
    else:
        config_path = Path(config_path).resolve()

    if not config_path.exists():
        print(f"ERROR: Config file not found at {config_path}")
        print("Create a config.ini with:")
        print("  [crossref]")
        print("  email = yourreal@email.com")
        print("  delay = 1")
        print("  save_every = 50")
        print("  max_retries = 3")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.read(config_path)
    email = config.get("crossref", "email", fallback=None)
    if not email or email.strip() == "your_email@example.com":
        print(f"ERROR: Please set your real email in {config_path}")
        sys.exit(1)

    return {
        "email": email.strip(),
        "delay": config.getfloat("crossref", "delay", fallback=1),
        "save_every": config.getint("crossref", "save_every", fallback=50),
        "max_retries": config.getint("crossref", "max_retries", fallback=3),
    }


# ── Cache ───────────────────────────────────────────────────────────────

def load_cache(cache_path):
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"  Loaded cache: {len(cache)} entries from {cache_path.name}")
        return cache
    return {}


def save_cache(cache, cache_path):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── Name matching ───────────────────────────────────────────────────────

def normalise_for_match(name):
    name = name.lower()
    name = name.replace('.', '').replace(',', '').replace('-', ' ')
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def name_tokens(name):
    return set(normalise_for_match(name).split())


def tokens_compatible(tokens_a, tokens_b):
    smaller, larger = (tokens_a, tokens_b) if len(tokens_a) <= len(tokens_b) else (tokens_b, tokens_a)
    for s in smaller:
        matched = any(
            s == l or (len(s) == 1 and l.startswith(s)) or (len(l) == 1 and s.startswith(l))
            for l in larger
        )
        if not matched:
            return False
    return True


def build_name_variants(author_name, variants_str):
    names = {author_name}
    if variants_str:
        for v in variants_str.split(';'):
            v = v.strip()
            if v:
                names.add(v)
    return names


# ── Crossref fetch ──────────────────────────────────────────────────────

def api_get_with_retry(url, session, params=None, max_retries=3, label=""):
    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, params=params, timeout=30)
            if resp.status_code == 404:
                return None
            if resp.status_code == 429:
                wait = min(2 ** attempt * 5, 60)
                print(f"    Rate limited. Waiting {wait}s (retry {attempt}/{max_retries})...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                wait = 2 ** attempt
                print(f"    {label} Error: {e}. Retrying in {wait}s ({attempt}/{max_retries})...")
                time.sleep(wait)
            else:
                print(f"    {label} Failed after {max_retries} retries: {e}")
                return None
    return None


def fetch_crossref_authors(doi, session, email, max_retries=3):
    url = f"{CROSSREF_API}{doi}"
    params = {"mailto": email}
    data = api_get_with_retry(url, session, params, max_retries, label="[Crossref]")
    if data is None:
        return []
    authors = data.get("message", {}).get("author", [])
    return [
        {
            "given": a.get("given", ""),
            "family": a.get("family", ""),
            "name": a.get("name", ""),
            "affiliation": a.get("affiliation", []),
        }
        for a in authors
    ]


def find_affiliations_crossref(target_names, crossref_authors):
    for cr_author in crossref_authors:
        given = cr_author.get("given", "")
        family = cr_author.get("family", "")
        cr_full = f"{given} {family}".strip()
        if not cr_full:
            cr_full = cr_author.get("name", "")
        if not cr_full:
            continue
        cr_tokens = name_tokens(cr_full)
        for candidate in target_names:
            if tokens_compatible(name_tokens(candidate), cr_tokens):
                affs = cr_author.get("affiliation", [])
                return [a.get("name", "") for a in affs if a.get("name")]
    return []


# ── OpenAlex fetch ──────────────────────────────────────────────────────

def fetch_openalex_authorships(doi, session, email, max_retries=3):
    url = f"{OPENALEX_API}{doi}"
    params = {"mailto": email}
    data = api_get_with_retry(url, session, params, max_retries, label="[OpenAlex]")
    if data is None:
        return []
    return data.get("authorships", [])


def find_affiliations_openalex(target_names, openalex_authorships):
    for authorship in openalex_authorships:
        oa_name = authorship.get("author", {}).get("display_name", "")
        raw_name = authorship.get("raw_author_name", "")
        if not oa_name and not raw_name:
            continue

        oa_names_to_check = set()
        if oa_name:
            oa_names_to_check.add(oa_name)
        if raw_name:
            oa_names_to_check.add(raw_name)

        for oa_n in oa_names_to_check:
            oa_tokens = name_tokens(oa_n)
            for candidate in target_names:
                if tokens_compatible(name_tokens(candidate), oa_tokens):
                    # Extract institution names
                    institutions = authorship.get("institutions", [])
                    affs = [inst.get("display_name", "") for inst in institutions if inst.get("display_name")]
                    if not affs:
                        # Try raw_affiliation_strings as fallback
                        raw_affs = authorship.get("raw_affiliation_strings", [])
                        affs = [r for r in raw_affs if r]
                    return affs
    return []


# ── Main ────────────────────────────────────────────────────────────────

def main(input_file, output_file=None, config_path=None):
    cfg = load_config(config_path)
    print(f"Config: email={cfg['email']}, delay={cfg['delay']}s, "
          f"save_every={cfg['save_every']}, max_retries={cfg['max_retries']}")

    # Resolve paths
    input_path = Path(input_file)
    if not input_path.is_absolute():
        if not input_path.exists():
            fallback = SCRIPT_DIR / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_with_affiliations.xlsx"
    else:
        output_path = Path(output_file).resolve()

    # Cache paths
    crossref_cache_path = output_path.parent / f"{input_path.stem}_crossref_doi_cache.json"
    openalex_cache_path = output_path.parent / f"{input_path.stem}_openalex_doi_cache.json"
    author_cache_path = output_path.parent / f"{input_path.stem}_affiliation_cache.json"

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    crossref_cache = load_cache(crossref_cache_path)
    openalex_cache = load_cache(openalex_cache_path)
    author_cache = load_cache(author_cache_path)

    # --- Read input ---
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    name_col = headers.get("Author_Name")
    variants_col = headers.get("Name_Variants")
    doi_count_col = headers.get("DOI_Count")
    dois_col = headers.get("DOIs")

    if name_col is None or dois_col is None:
        print("ERROR: Input must have 'Author_Name' and 'DOIs' columns.")
        sys.exit(1)

    # Collect authors
    authors = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        variants = ws.cell(row=row, column=variants_col).value if variants_col else ""
        doi_count = ws.cell(row=row, column=doi_count_col).value if doi_count_col else 0
        dois_str = ws.cell(row=row, column=dois_col).value
        if not name or not dois_str:
            continue
        dois = [d.strip() for d in str(dois_str).split(';') if d.strip()]
        authors.append({
            "row": row,
            "name": str(name).strip(),
            "variants": str(variants).strip() if variants else "",
            "doi_count": doi_count,
            "dois": dois,
        })

    # Collect unique DOIs
    all_dois = set()
    for a in authors:
        all_dois.update(a["dois"])

    print(f"\nFound {len(authors)} authors referencing {len(all_dois)} unique DOIs.")

    # --- Phase 1: Fetch Crossref metadata ---
    session = requests.Session()
    session.headers.update({"User-Agent": f"AffiliationFetch/1.0 (mailto:{cfg['email']})"})

    cr_to_fetch = [d for d in sorted(all_dois) if d not in crossref_cache]
    print(f"\n--- Phase 1: Crossref ---")
    print(f"{len(all_dois) - len(cr_to_fetch)} cached, {len(cr_to_fetch)} to fetch.")

    fetched = 0
    try:
        for doi in cr_to_fetch:
            fetched += 1
            print(f"  [Crossref {fetched}/{len(cr_to_fetch)}] {doi}")
            result = fetch_crossref_authors(doi, session, cfg["email"], cfg["max_retries"])
            crossref_cache[doi] = result

            if fetched % cfg["save_every"] == 0:
                save_cache(crossref_cache, crossref_cache_path)
                print(f"    >> Crossref cache saved ({fetched}/{len(cr_to_fetch)})")
            if fetched < len(cr_to_fetch):
                time.sleep(cfg["delay"])
    except KeyboardInterrupt:
        print(f"\n>> Interrupted! Saving Crossref cache ({fetched} fetched)...")
        save_cache(crossref_cache, crossref_cache_path)
        print("Re-run to continue.")
        sys.exit(0)

    save_cache(crossref_cache, crossref_cache_path)

    # --- Resolve Crossref affiliations ---
    authors_needing_openalex = []

    for a in authors:
        cache_key = a["name"]
        if cache_key in author_cache:
            a["affiliations"] = author_cache[cache_key]["affiliations"]
            a["source"] = author_cache[cache_key]["source"]
            continue

        target_names = build_name_variants(a["name"], a["variants"])
        all_affiliations = set()

        for doi in a["dois"]:
            cr_authors = crossref_cache.get(doi, [])
            affs = find_affiliations_crossref(target_names, cr_authors)
            all_affiliations.update(affs)

        if all_affiliations:
            a["affiliations"] = sorted(all_affiliations)
            a["source"] = "Crossref"
            author_cache[cache_key] = {"affiliations": a["affiliations"], "source": "Crossref"}
        else:
            authors_needing_openalex.append(a)

    save_cache(author_cache, author_cache_path)

    cr_found = sum(1 for a in authors if a.get("source") == "Crossref")
    print(f"\nCrossref resolved affiliations for {cr_found}/{len(authors)} authors.")
    print(f"{len(authors_needing_openalex)} authors need OpenAlex fallback.")

    # --- Phase 2: OpenAlex fallback ---
    if authors_needing_openalex:
        # Collect DOIs that need OpenAlex lookup
        openalex_dois = set()
        for a in authors_needing_openalex:
            openalex_dois.update(a["dois"])

        oa_to_fetch = [d for d in sorted(openalex_dois) if d not in openalex_cache]
        print(f"\n--- Phase 2: OpenAlex ---")
        print(f"{len(openalex_dois) - len(oa_to_fetch)} cached, {len(oa_to_fetch)} to fetch.")

        fetched = 0
        try:
            for doi in oa_to_fetch:
                fetched += 1
                print(f"  [OpenAlex {fetched}/{len(oa_to_fetch)}] {doi}")
                result = fetch_openalex_authorships(doi, session, cfg["email"], cfg["max_retries"])
                # Store compact version
                openalex_cache[doi] = [
                    {
                        "author": auth.get("author", {}),
                        "raw_author_name": auth.get("raw_author_name", ""),
                        "institutions": auth.get("institutions", []),
                        "raw_affiliation_strings": auth.get("raw_affiliation_strings", []),
                    }
                    for auth in result
                ] if result else []

                if fetched % cfg["save_every"] == 0:
                    save_cache(openalex_cache, openalex_cache_path)
                    print(f"    >> OpenAlex cache saved ({fetched}/{len(oa_to_fetch)})")
                if fetched < len(oa_to_fetch):
                    time.sleep(cfg["delay"])
        except KeyboardInterrupt:
            print(f"\n>> Interrupted! Saving OpenAlex cache ({fetched} fetched)...")
            save_cache(openalex_cache, openalex_cache_path)
            print("Re-run to continue.")
            sys.exit(0)

        save_cache(openalex_cache, openalex_cache_path)

        # Resolve OpenAlex affiliations
        for a in authors_needing_openalex:
            cache_key = a["name"]
            if cache_key in author_cache:
                a["affiliations"] = author_cache[cache_key]["affiliations"]
                a["source"] = author_cache[cache_key]["source"]
                continue

            target_names = build_name_variants(a["name"], a["variants"])
            all_affiliations = set()

            for doi in a["dois"]:
                oa_authorships = openalex_cache.get(doi, [])
                affs = find_affiliations_openalex(target_names, oa_authorships)
                all_affiliations.update(affs)

            if all_affiliations:
                a["affiliations"] = sorted(all_affiliations)
                a["source"] = "OpenAlex"
            else:
                a["affiliations"] = ["[No affiliation found]"]
                a["source"] = "None"

            author_cache[cache_key] = {"affiliations": a["affiliations"], "source": a["source"]}

        save_cache(author_cache, author_cache_path)

    # Fill in any remaining authors not yet processed
    for a in authors:
        if "affiliations" not in a:
            a["affiliations"] = ["[No affiliation found]"]
            a["source"] = "None"

    # --- Stats ---
    from_crossref = sum(1 for a in authors if a.get("source") == "Crossref")
    from_openalex = sum(1 for a in authors if a.get("source") == "OpenAlex")
    not_found = sum(1 for a in authors if a.get("source") == "None")

    print(f"\n--- Summary ---")
    print(f"  Crossref:  {from_crossref} authors")
    print(f"  OpenAlex:  {from_openalex} authors")
    print(f"  Not found: {not_found} authors")
    print(f"  Total:     {len(authors)} authors")

    # --- Write output ---
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Authors with Affiliations"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center")

    col_headers = [
        "Author_Name", "Name_Variants", "Affiliations",
        "Affiliation_Source", "DOI_Count", "DOIs"
    ]
    for col_idx, header in enumerate(col_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data_font = Font(name="Arial")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_idx, a in enumerate(authors, 2):
        out_ws.cell(row=row_idx, column=1, value=a["name"]).font = data_font
        out_ws.cell(row=row_idx, column=2, value=a["variants"]).font = data_font

        affs_str = "; ".join(a["affiliations"])
        cell = out_ws.cell(row=row_idx, column=3, value=affs_str)
        cell.font = data_font
        cell.alignment = wrap_align

        out_ws.cell(row=row_idx, column=4, value=a.get("source", "")).font = data_font
        out_ws.cell(row=row_idx, column=5, value=a["doi_count"]).font = data_font
        out_ws.cell(row=row_idx, column=6, value="; ".join(a["dois"])).font = data_font

    out_ws.column_dimensions['A'].width = 30
    out_ws.column_dimensions['B'].width = 40
    out_ws.column_dimensions['C'].width = 70
    out_ws.column_dimensions['D'].width = 18
    out_ws.column_dimensions['E'].width = 12
    out_ws.column_dimensions['F'].width = 70
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = f"A1:F{out_ws.max_row}"

    out_wb.save(output_path)
    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python fetch_affiliations.py <unique_authors.xlsx> [output.xlsx] [config.ini]")
        sys.exit(1)
    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    cfg_arg = sys.argv[3] if len(sys.argv) > 3 else None
    main(in_arg, out_arg, cfg_arg)