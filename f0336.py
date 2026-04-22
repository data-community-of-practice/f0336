"""
Affiliation Fetch (Crossref -> ORCID -> LLM)
===============================================
Reads the unique authors Excel file and fetches each author's affiliated
organisation(s) using a three-tier strategy:

  Tier 1 - Crossref: per DOI, extract author affiliations (most structured).
           Also extracts ORCID IDs when available.
  Tier 2 - ORCID: if author has an ORCID ID (from Crossref), fetch their
           employment affiliations directly from their public profile.
           NO credentials needed — public data is freely accessible.
  Tier 3 - LLM: for remaining authors, use Claude to infer likely
           affiliation (cached, temperature=0).

Setup:
  config.ini with:
    [crossref]
    email = yourreal@email.com
    delay = 1
    save_every = 50
    max_retries = 3

    [anthropic]                      # OPTIONAL - only for LLM fallback
    api_key = sk-ant-...

Usage:
  python fetch_affiliations.py <unique_authors.xlsx> [output.xlsx] [config.ini]
"""

import sys
import re
import json
import time
import configparser
from pathlib import Path
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

CROSSREF_API = "https://api.crossref.org/works/"
ORCID_API = "https://pub.orcid.org/v3.0"
ANTHROPIC_API = "https://api.anthropic.com/v1/messages"
SCRIPT_DIR = Path(__file__).resolve().parent


# ── Config ──────────────────────────────────────────────────────────────

def load_config(config_path=None):
    if config_path is None:
        config_path = SCRIPT_DIR / "config.ini"
    else:
        config_path = Path(config_path).resolve()

    if not config_path.exists():
        print(f"ERROR: Config file not found at {config_path}")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.read(config_path)
    email = config.get("crossref", "email", fallback=None)
    if not email or email.strip() == "your_email@example.com":
        print(f"ERROR: Please set your real email in {config_path}")
        sys.exit(1)

    return {
        "email": email.strip(),
        "delay": config.getfloat("crossref", "delay", fallback=1),
        "save_every": config.getint("crossref", "save_every", fallback=50),
        "max_retries": config.getint("crossref", "max_retries", fallback=3),
        "api_key": config.get("anthropic", "api_key", fallback="").strip(),
    }


# ── Cache ───────────────────────────────────────────────────────────────

def load_cache(cache_path):
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_cache(cache, cache_path):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── Name matching ───────────────────────────────────────────────────────

def normalise_for_match(name):
    name = name.lower().replace('.', '').replace(',', '').replace('-', ' ')
    return re.sub(r'\s+', ' ', name).strip()

def name_tokens(name):
    return set(normalise_for_match(name).split())

def tokens_compatible(tokens_a, tokens_b):
    smaller, larger = (tokens_a, tokens_b) if len(tokens_a) <= len(tokens_b) else (tokens_b, tokens_a)
    for s in smaller:
        if not any(s == l or (len(s) == 1 and l.startswith(s)) or (len(l) == 1 and s.startswith(l)) for l in larger):
            return False
    return True

def build_name_variants(author_name, variants_str):
    names = {author_name}
    if variants_str:
        for v in variants_str.split(';'):
            v = v.strip()
            if v:
                names.add(v)
    return names


# ── API helper ──────────────────────────────────────────────────────────

def api_get(url, session, params=None, headers=None, max_retries=3):
    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, params=params, headers=headers, timeout=30)
            if resp.status_code == 404:
                return None
            if resp.status_code == 429:
                time.sleep(min(2 ** attempt * 5, 60))
                continue
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                return None
    return None


# ── Tier 1: Crossref ───────────────────────────────────────────────────

def fetch_crossref_authors(doi, session, email, max_retries=3):
    data = api_get(f"{CROSSREF_API}{doi}", session, params={"mailto": email}, max_retries=max_retries)
    if data is None:
        return []
    return [
        {
            "given": a.get("given", ""),
            "family": a.get("family", ""),
            "name": a.get("name", ""),
            "orcid": a.get("ORCID", "").replace("http://orcid.org/", "").replace("https://orcid.org/", ""),
            "affiliation": a.get("affiliation", []),
        }
        for a in data.get("message", {}).get("author", [])
    ]

def find_affiliations_crossref(target_names, crossref_authors):
    for cr in crossref_authors:
        cr_full = f"{cr.get('given', '')} {cr.get('family', '')}".strip() or cr.get("name", "")
        if not cr_full:
            continue
        for candidate in target_names:
            if tokens_compatible(name_tokens(candidate), name_tokens(cr_full)):
                affs = [a.get("name", "") for a in cr.get("affiliation", []) if a.get("name")]
                return affs, cr.get("orcid", "")
    return [], ""


# ── Tier 2: ORCID (direct lookup, no credentials) ─────────────────────

def fetch_orcid_employments(orcid_id, session, max_retries=3):
    """
    Fetch employment affiliations from a public ORCID profile.
    No credentials needed — just the ORCID ID.
    """
    url = f"{ORCID_API}/{orcid_id}/employments"
    headers = {"Accept": "application/json"}

    data = api_get(url, session, headers=headers, max_retries=max_retries)
    if data is None:
        return []

    affs = []

    # ORCID v3 nests employments in affiliation-group -> summaries
    for group in data.get("affiliation-group", []):
        for summary in group.get("summaries", []):
            emp = summary.get("employment-summary", {})
            org_name = emp.get("organization", {}).get("name", "")
            if org_name and org_name not in affs:
                affs.append(org_name)

    # Fallback: some API responses use top-level employment-summary
    if not affs:
        for item in data.get("employment-summary", []):
            org_name = item.get("organization", {}).get("name", "")
            if org_name and org_name not in affs:
                affs.append(org_name)

    return affs


# ── Tier 3: LLM ────────────────────────────────────────────────────────

LLM_AFFILIATION_PROMPT = """You are extracting the most likely current institutional affiliation for an academic researcher.

Given an author name and any available context (name variants, DOIs they published), determine their most likely primary institutional affiliation.

Rules:
- Return the name of the INSTITUTION only (university, hospital, research institute, company).
- Do NOT include department names, faculty names, or addresses.
- Assign a confidence score from 0-100 for the affiliation
- If you cannot determine the affiliation with confidence score greaten than 70, return UNKNOWN.

Respond with ONLY a JSON object on a single line, no markdown:
{"affiliation": "Institution Name", "confidence": "0-100"}

If unknown:
{"affiliation": "UNKNOWN", "confidence": "20"}"""


def infer_affiliation_llm(author_name, variants, dois_str, api_key, session, cache, max_retries=3):
    cache_key = author_name.lower().strip()
    if cache_key in cache:
        return cache[cache_key].get("affiliation", ""), cache[cache_key].get("confidence", "")

    if not api_key:
        return "", ""

    user_msg = f'Author: "{author_name}"'
    if variants:
        user_msg += f'\nName variants: {variants}'
    if dois_str:
        user_msg += f'\nPublished DOIs: {dois_str[:500]}'

    payload = {
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 100,
        "temperature": 0,
        "system": LLM_AFFILIATION_PROMPT,
        "messages": [{"role": "user", "content": user_msg}],
    }
    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
    }

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.post(ANTHROPIC_API, json=payload, headers=headers, timeout=60)
            if resp.status_code == 429:
                time.sleep(min(2 ** attempt * 5, 60))
                continue
            resp.raise_for_status()
            data = resp.json()

            text = "".join(b.get("text", "") for b in data.get("content", []) if b.get("type") == "text").strip()
            text = text.replace("```json", "").replace("```", "").strip()
            parsed = json.loads(text)

            aff = parsed.get("affiliation", "").strip()
            conf = parsed.get("confidence", "low").strip()
            if aff == "UNKNOWN":
                aff = ""

            cache[cache_key] = {"affiliation": aff, "confidence": conf}
            return aff, conf

        except (requests.exceptions.RequestException, json.JSONDecodeError):
            if attempt < max_retries:
                time.sleep(2 ** attempt)

    cache[cache_key] = {"affiliation": "", "confidence": ""}
    return "", ""


# ── Main ────────────────────────────────────────────────────────────────

def main(input_file, output_file=None, config_path=None):
    cfg = load_config(config_path)
    has_llm = bool(cfg["api_key"])
    print(f"Config: email={cfg['email']}")
    print(f"  ORCID direct lookup: always enabled (no key needed)")
    print(f"  LLM tier:            {'enabled' if has_llm else 'disabled'}")

    input_path = Path(input_file)
    if not input_path.is_absolute():
        if not input_path.exists():
            fallback = SCRIPT_DIR / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_with_affiliations.xlsx"
    else:
        output_path = Path(output_file).resolve()

    crossref_cache_path = output_path.parent / f"{input_path.stem}_crossref_doi_cache.json"
    orcid_cache_path = output_path.parent / f"{input_path.stem}_orcid_cache.json"
    llm_cache_path = output_path.parent / f"{input_path.stem}_llm_aff_cache.json"
    author_cache_path = output_path.parent / f"{input_path.stem}_affiliation_cache.json"

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    crossref_cache = load_cache(crossref_cache_path)
    orcid_cache = load_cache(orcid_cache_path)
    llm_cache = load_cache(llm_cache_path)
    author_cache = load_cache(author_cache_path)

    for name, c in [("Crossref DOI", crossref_cache), ("ORCID", orcid_cache),
                     ("LLM", llm_cache), ("Author", author_cache)]:
        if c:
            print(f"  {name} cache: {len(c)} entries")

    # --- Read input ---
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    hdr = {cell.value: cell.column for cell in ws[1]}
    name_col = hdr.get("Author_Name")
    variants_col = hdr.get("Name_Variants")
    doi_count_col = hdr.get("DOI_Count")
    dois_col = hdr.get("DOIs")

    if name_col is None or dois_col is None:
        print("ERROR: Input must have 'Author_Name' and 'DOIs' columns.")
        sys.exit(1)

    authors = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        variants = ws.cell(row=row, column=variants_col).value if variants_col else ""
        doi_count = ws.cell(row=row, column=doi_count_col).value if doi_count_col else 0
        dois_str = ws.cell(row=row, column=dois_col).value
        if not name or not dois_str:
            continue
        dois = [d.strip() for d in str(dois_str).split(';') if d.strip()]
        authors.append({
            "name": str(name).strip(),
            "variants": str(variants).strip() if variants else "",
            "doi_count": doi_count,
            "dois": dois,
        })

    all_dois = set()
    for a in authors:
        all_dois.update(a["dois"])

    print(f"\nFound {len(authors)} authors referencing {len(all_dois)} unique DOIs.")

    session = requests.Session()
    session.headers.update({"User-Agent": f"AffiliationFetch/2.0 (mailto:{cfg['email']})"})

    # ── Tier 1: Crossref ────────────────────────────────────────────
    print(f"\n--- Tier 1: Crossref ---")
    cr_to_fetch = [d for d in sorted(all_dois) if d not in crossref_cache]
    print(f"  {len(all_dois) - len(cr_to_fetch)} cached, {len(cr_to_fetch)} to fetch.")

    fetched = 0
    try:
        for doi in cr_to_fetch:
            fetched += 1
            display = doi[:60].encode("ascii", errors="replace").decode("ascii")
            print(f"  [{fetched}/{len(cr_to_fetch)}] {display}")
            crossref_cache[doi] = fetch_crossref_authors(doi, session, cfg["email"], cfg["max_retries"])
            if fetched % cfg["save_every"] == 0:
                save_cache(crossref_cache, crossref_cache_path)
            if fetched < len(cr_to_fetch):
                time.sleep(cfg["delay"])
    except KeyboardInterrupt:
        print(f"\n>> Interrupted! Saving...")
        save_cache(crossref_cache, crossref_cache_path)
        sys.exit(0)

    save_cache(crossref_cache, crossref_cache_path)

    # Resolve Crossref affiliations + collect ORCID IDs
    # For each author: check DOIs one by one, stop as soon as affiliation is found.
    # Still scan remaining DOIs for ORCID ID if we haven't found one yet.
    tier2_needed = []
    for a in authors:
        cache_key = a["name"]
        if cache_key in author_cache:
            a["affiliations"] = author_cache[cache_key]["affiliations"]
            a["source"] = author_cache[cache_key]["source"]
            continue

        target_names = build_name_variants(a["name"], a["variants"])
        found_affs = []
        found_orcid = ""

        for doi in a["dois"]:
            cr_authors = crossref_cache.get(doi, [])
            affs, orcid = find_affiliations_crossref(target_names, cr_authors)

            if affs and not found_affs:
                found_affs = affs

            if orcid and not found_orcid:
                found_orcid = orcid

            # Stop early: we have both affiliation and ORCID
            if found_affs and found_orcid:
                break

            # Stop if we have affiliation (ORCID is nice-to-have)
            if found_affs:
                break

        if found_affs:
            a["affiliations"] = sorted(set(found_affs))
            a["source"] = "Crossref"
            a["orcid"] = found_orcid
            author_cache[cache_key] = {"affiliations": a["affiliations"], "source": "Crossref"}
        else:
            a["orcid"] = found_orcid
            tier2_needed.append(a)

    save_cache(author_cache, author_cache_path)
    cr_found = sum(1 for a in authors if a.get("source") == "Crossref")
    orcid_ids_found = sum(1 for a in tier2_needed if a.get("orcid"))
    print(f"  Resolved: {cr_found}/{len(authors)} authors")
    print(f"  ORCID IDs available for Tier 2: {orcid_ids_found}/{len(tier2_needed)} remaining authors")

    # ── Tier 2: ORCID (direct lookup only) ─────────────────────────
    print(f"\n--- Tier 2: ORCID (direct lookup) ---")

    # Split into authors with ORCID IDs and those without
    has_orcid = [a for a in tier2_needed if a.get("orcid")]
    no_orcid = [a for a in tier2_needed if not a.get("orcid")]

    print(f"  {len(has_orcid)} authors with ORCID IDs to look up.")
    print(f"  {len(no_orcid)} authors without ORCID IDs (skipping to Tier 3).")

    tier3_needed = list(no_orcid)  # these go straight to LLM
    orcid_found = 0

    try:
        for i, a in enumerate(has_orcid):
            cache_key = a["name"]
            orcid_id = a["orcid"]
            display = a["name"][:45].encode("ascii", errors="replace").decode("ascii")

            # Check cache
            if cache_key in orcid_cache:
                cached_affs = orcid_cache[cache_key].get("affiliations", [])
                if cached_affs:
                    a["affiliations"] = cached_affs
                    a["source"] = "ORCID"
                    author_cache[cache_key] = {"affiliations": cached_affs, "source": "ORCID"}
                    orcid_found += 1
                else:
                    tier3_needed.append(a)
                continue

            print(f"  [{i+1}/{len(has_orcid)}] {display} ({orcid_id})")
            affs = fetch_orcid_employments(orcid_id, session, cfg["max_retries"])

            orcid_cache[cache_key] = {"affiliations": affs, "orcid_id": orcid_id}

            if affs:
                a["affiliations"] = affs
                a["source"] = "ORCID"
                author_cache[cache_key] = {"affiliations": affs, "source": "ORCID"}
                orcid_found += 1
            else:
                tier3_needed.append(a)

            if (i + 1) % cfg["save_every"] == 0:
                save_cache(orcid_cache, orcid_cache_path)
                save_cache(author_cache, author_cache_path)

            time.sleep(cfg["delay"])

    except KeyboardInterrupt:
        print(f"\n>> Interrupted! Saving...")
        save_cache(orcid_cache, orcid_cache_path)
        save_cache(author_cache, author_cache_path)
        sys.exit(0)

    save_cache(orcid_cache, orcid_cache_path)
    save_cache(author_cache, author_cache_path)
    print(f"  Resolved: {orcid_found}/{len(has_orcid)} authors")

    # ── Tier 3: LLM ────────────────────────────────────────────────
    print(f"\n--- Tier 3: LLM ---")

    if not has_llm:
        print(f"  Skipped (no API key). {len(tier3_needed)} authors unresolved.")
        for a in tier3_needed:
            a["affiliations"] = ["[No affiliation found]"]
            a["source"] = "None"
    else:
        print(f"  {len(tier3_needed)} authors to infer.")
        llm_found = 0
        try:
            for i, a in enumerate(tier3_needed):
                display = a["name"][:50].encode("ascii", errors="replace").decode("ascii")
                print(f"  [{i+1}/{len(tier3_needed)}] {display}")

                aff, conf = infer_affiliation_llm(
                    a["name"], a["variants"], "; ".join(a["dois"]),
                    cfg["api_key"], session, llm_cache, cfg["max_retries"]
                )

                if aff:
                    a["affiliations"] = [aff]
                    a["source"] = f"LLM ({conf})"
                    author_cache[a["name"]] = {"affiliations": [aff], "source": f"LLM ({conf})"}
                    llm_found += 1
                else:
                    a["affiliations"] = ["[No affiliation found]"]
                    a["source"] = "None"

                if (i + 1) % cfg["save_every"] == 0:
                    save_cache(llm_cache, llm_cache_path)

                time.sleep(cfg["delay"])

        except KeyboardInterrupt:
            print(f"\n>> Interrupted! Saving...")

        save_cache(llm_cache, llm_cache_path)
        save_cache(author_cache, author_cache_path)
        print(f"  Resolved: {llm_found}/{len(tier3_needed)} authors")

    # Fill remaining
    for a in authors:
        if "affiliations" not in a:
            a["affiliations"] = ["[No affiliation found]"]
            a["source"] = "None"

    # --- Stats ---
    from collections import Counter
    src_counts = Counter(a.get("source", "None") for a in authors)
    print(f"\n--- Summary ---")
    for src in sorted(src_counts.keys()):
        print(f"  {src}: {src_counts[src]}")
    print(f"  Total: {len(authors)}")

    # --- Write output ---
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Authors with Affiliations"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center")

    col_headers = [
        "Author_Name", "Name_Variants", "Affiliations",
        "Affiliation_Source", "DOI_Count", "DOIs"
    ]
    for col_idx, header in enumerate(col_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data_font = Font(name="Arial")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_idx, a in enumerate(authors, 2):
        out_ws.cell(row=row_idx, column=1, value=a["name"]).font = data_font
        out_ws.cell(row=row_idx, column=2, value=a["variants"]).font = data_font
        cell = out_ws.cell(row=row_idx, column=3, value="; ".join(a["affiliations"]))
        cell.font = data_font
        cell.alignment = wrap_align
        out_ws.cell(row=row_idx, column=4, value=a.get("source", "")).font = data_font
        out_ws.cell(row=row_idx, column=5, value=a["doi_count"]).font = data_font
        out_ws.cell(row=row_idx, column=6, value="; ".join(a["dois"])).font = data_font

    out_ws.column_dimensions['A'].width = 30
    out_ws.column_dimensions['B'].width = 40
    out_ws.column_dimensions['C'].width = 70
    out_ws.column_dimensions['D'].width = 18
    out_ws.column_dimensions['E'].width = 12
    out_ws.column_dimensions['F'].width = 70
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = f"A1:F{out_ws.max_row}"

    out_wb.save(output_path)
    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python fetch_affiliations.py <unique_authors.xlsx> [output.xlsx] [config.ini]")
        sys.exit(1)
    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    cfg_arg = sys.argv[3] if len(sys.argv) > 3 else None
    main(in_arg, out_arg, cfg_arg)