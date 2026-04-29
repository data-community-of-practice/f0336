"""
Crossref Author Metadata Fetcher
==================================
Reads an Excel file with a 'Publication_DOI' column, queries the Crossref
Works API for each unique DOI, and produces a structured JSON file with
author metadata (names, ORCID IDs, affiliations with ROR and place).

Reproducibility guarantees:
  - No inference, correction, or invention of author names.
  - No web search or sources other than Crossref.
  - No paraphrasing of metadata.
  - Crossref author order preserved exactly as returned.
  - Unicode characters preserved.
  - Deterministic processing: same input always produces same output.
  - JSON cache prevents redundant API calls across runs.

Usage:
  python f0334.py <input.xlsx> <output.json> [config.ini]

Example:
  python f0334.py Publications_with_high_confidence.xlsx Crossref_AuthorMetadata.json config.ini
"""

import sys
import json
import time
import configparser
from pathlib import Path
from collections import OrderedDict
import requests
import openpyxl

CROSSREF_API = "https://api.crossref.org/works/"
SCRIPT_DIR = Path(__file__).resolve().parent


# ── Config ──────────────────────────────────────────────────────────────

def load_config(config_path=None):
    if config_path is None:
        config_path = SCRIPT_DIR / "config.ini"
    else:
        config_path = Path(config_path).resolve()

    if not config_path.exists():
        print(f"ERROR: Config file not found at {config_path}")
        print("Create a config.ini with:")
        print("  [crossref]")
        print("  email = yourreal@email.com")
        print("  delay = 1")
        print("  save_every = 50")
        print("  max_retries = 3")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.read(config_path)
    email = config.get("crossref", "email", fallback=None)
    if not email or email.strip() == "your_email@example.com":
        print(f"ERROR: Please set your real email in {config_path}")
        sys.exit(1)

    return {
        "email": email.strip(),
        "delay": config.getfloat("crossref", "delay", fallback=1),
        "save_every": config.getint("crossref", "save_every", fallback=50),
        "max_retries": config.getint("crossref", "max_retries", fallback=3),
    }


# ── Cache ───────────────────────────────────────────────────────────────

def load_cache(cache_path):
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"Loaded cache with {len(cache)} DOIs from {cache_path.name}")
        return cache
    return {}


def save_cache(cache, cache_path):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── Crossref fetch ─────────────────────────────────────────────────────

def fetch_crossref_record(doi, session, email, max_retries=3):
    """
    Fetch the raw Crossref record for a DOI.
    Returns the message dict or None (for 404).
    """
    url = f"{CROSSREF_API}{doi}"
    params = {"mailto": email}

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(url, params=params, timeout=30)

            if resp.status_code == 404:
                return None

            if resp.status_code == 429:
                wait = min(2 ** attempt * 5, 60)
                print(f"    Rate limited. Waiting {wait}s (retry {attempt}/{max_retries})...")
                time.sleep(wait)
                continue

            resp.raise_for_status()
            data = resp.json()
            return data.get("message", {})

        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                wait = 2 ** attempt
                print(f"    Error: {e}. Retrying in {wait}s ({attempt}/{max_retries})...")
                time.sleep(wait)
            else:
                print(f"    Failed after {max_retries} retries: {e}")
                return {"_error": str(e)}

    return {"_error": f"Failed after {max_retries} retries"}


# ── Extract structured author metadata ─────────────────────────────────

def extract_author_metadata(message, doi):
    """
    Extract structured author metadata from a Crossref message.
    No inference, no correction — only what Crossref returns.
    """
    if message is None:
        return {
            "doi": doi,
            "error": "DOI not found in Crossref",
            "authors": [],
        }

    if "_error" in message:
        return {
            "doi": doi,
            "error": message["_error"],
            "authors": [],
        }

    # Extract title
    titles = message.get("title", [])
    title = titles[0] if titles else ""

    # Extract authors — preserve order exactly as returned
    author_list = message.get("author", [])
    authors = []

    for a in author_list:
        # Extract name fields exactly as Crossref provides
        given = a.get("given", "")
        family = a.get("family", "")

        # ORCID: normalise by stripping URL prefix
        orcid = a.get("ORCID", "")
        if orcid:
            orcid = orcid.replace("http://orcid.org/", "").replace("https://orcid.org/", "")

        # Extract affiliations with ROR and place
        affiliations = []
        for aff in a.get("affiliation", []):
            aff_entry = {}

            aff_name = aff.get("name", "")
            if aff_name:
                aff_entry["name"] = aff_name

            # Extract ROR identifier(s)
            ror_ids = []
            for id_entry in aff.get("id", []):
                if id_entry.get("id-type", "").upper() == "ROR":
                    ror_id = id_entry.get("id", "")
                    if ror_id:
                        ror_ids.append(ror_id)

            if ror_ids:
                aff_entry["ror_id"] = ror_ids[0]
                if len(ror_ids) > 1:
                    aff_entry["ror_ids_all"] = ror_ids

            # Extract place
            places = aff.get("place", [])
            if places:
                aff_entry["place"] = places[0] if len(places) == 1 else places

            if aff_entry:
                affiliations.append(aff_entry)

        author_entry = {
            "given": given,
            "family": family,
            "orcid": orcid,
            "affiliations": affiliations,
        }

        # Include organizational name if no given/family
        org_name = a.get("name", "")
        if org_name and not given and not family:
            author_entry["name"] = org_name

        authors.append(author_entry)

    result = {
        "doi": doi,
        "title": title,
        "authors": authors,
    }

    return result


# ── Read DOIs from Excel ───────────────────────────────────────────────

def read_dois_from_excel(input_path):
    """
    Read all non-empty DOI values from the Publication_DOI column.
    Remove duplicates using case-insensitive matching while preserving
    first occurrence order.
    """
    wb = openpyxl.load_workbook(input_path, read_only=True)
    ws = wb.active

    # Find the DOI column
    headers = {cell.value: cell.column for cell in ws[1]}
    doi_col = headers.get("Publication_DOI")
    if doi_col is None:
        print("ERROR: Could not find 'Publication_DOI' column in the header row.")
        sys.exit(1)

    # Collect DOIs preserving first occurrence order, case-insensitive dedup
    seen_lower = set()
    unique_dois = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        cell_value = row[doi_col - 1].value
        if cell_value is None:
            continue

        doi = str(cell_value).strip()
        if not doi:
            continue

        doi_lower = doi.lower()
        if doi_lower not in seen_lower:
            seen_lower.add(doi_lower)
            unique_dois.append(doi)

    wb.close()
    return unique_dois


# ── Main ────────────────────────────────────────────────────────────────

def main(input_file, output_file, config_path=None):
    cfg = load_config(config_path)
    print(f"Config: email={cfg['email']}, delay={cfg['delay']}s, "
          f"save_every={cfg['save_every']}, max_retries={cfg['max_retries']}")

    input_path = Path(input_file)
    if not input_path.is_absolute():
        if not input_path.exists():
            fallback = SCRIPT_DIR / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    output_path = Path(output_file).resolve() if Path(output_file).is_absolute() else Path(output_file)
    cache_path = output_path.parent / f"{output_path.stem}_cache.json"

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print(f"Cache:  {cache_path}")

    # Read DOIs
    unique_dois = read_dois_from_excel(input_path)
    print(f"\nFound {len(unique_dois)} unique DOIs.")

    # Load cache
    cache = load_cache(cache_path)
    cached_count = sum(1 for d in unique_dois if d in cache)
    to_fetch = len(unique_dois) - cached_count
    print(f"  Cached: {cached_count}, to fetch: {to_fetch}")

    # Fetch from Crossref
    if to_fetch > 0:
        session = requests.Session()
        session.headers.update({
            "User-Agent": f"CrossrefAuthorMetadata/1.0 (mailto:{cfg['email']})",
        })

        fetched = 0
        try:
            for doi in unique_dois:
                if doi in cache:
                    continue

                fetched += 1
                display = doi[:60].encode("ascii", errors="replace").decode("ascii")
                print(f"[{fetched}/{to_fetch}] {display}")

                record = fetch_crossref_record(doi, session, cfg["email"], cfg["max_retries"])

                # Store the raw message in cache (None for 404s)
                if record is None:
                    cache[doi] = None
                else:
                    # Store only the fields we need to keep cache size manageable
                    cache[doi] = {
                        "title": record.get("title", []),
                        "author": record.get("author", []),
                        "_error": record.get("_error"),
                    }

                if fetched % cfg["save_every"] == 0:
                    save_cache(cache, cache_path)
                    print(f"  >> Cache saved ({fetched}/{to_fetch})")

                if fetched < to_fetch:
                    time.sleep(cfg["delay"])

        except KeyboardInterrupt:
            print(f"\n>> Interrupted! Saving cache ({fetched} fetched)...")
            save_cache(cache, cache_path)
            print("Re-run to continue.")
            sys.exit(0)

        save_cache(cache, cache_path)

    # Build output JSON
    print(f"\nBuilding output...")
    results = []

    for doi in unique_dois:
        cached_record = cache.get(doi)

        if cached_record is None:
            # 404 — DOI not found
            results.append({
                "doi": doi,
                "error": "DOI not found in Crossref",
                "authors": [],
            })
        elif cached_record.get("_error"):
            results.append({
                "doi": doi,
                "error": cached_record["_error"],
                "authors": [],
            })
        else:
            # Reconstruct message format for extraction
            message = {
                "title": cached_record.get("title", []),
                "author": cached_record.get("author", []),
            }
            results.append(extract_author_metadata(message, doi))

    # Write output JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # Summary statistics
    total_dois = len(results)
    total_authors = 0
    authors_with_orcid = 0
    authors_with_affiliation = 0
    authors_without_orcid = 0
    dois_with_error = 0

    for entry in results:
        if "error" in entry:
            dois_with_error += 1
        for author in entry.get("authors", []):
            total_authors += 1
            if author.get("orcid"):
                authors_with_orcid += 1
            else:
                authors_without_orcid += 1
            if author.get("affiliations"):
                authors_with_affiliation += 1

    print(f"\n{'=' * 50}")
    print(f"SUMMARY")
    print(f"{'=' * 50}")
    print(f"  DOIs processed:           {total_dois}")
    print(f"  DOIs with errors/404:     {dois_with_error}")
    print(f"  Total authors:            {total_authors}")
    print(f"  Authors with ORCID:       {authors_with_orcid}")
    print(f"  Authors without ORCID:    {authors_without_orcid}")
    print(f"  Authors with affiliation: {authors_with_affiliation}")
    print(f"{'=' * 50}")
    print(f"\nOutput saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python f0334.py <input.xlsx> <output.json> [config.ini]")
        print("Example: python f0334.py Publications_with_high_confidence.xlsx Crossref_AuthorMetadata.json config.ini")
        sys.exit(1)

    input_arg = sys.argv[1]
    output_arg = sys.argv[2]
    config_arg = sys.argv[3] if len(sys.argv) > 3 else None
    main(input_arg, output_arg, config_arg)